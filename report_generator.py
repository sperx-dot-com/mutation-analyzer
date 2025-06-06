import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import MergedCell
import os

def generate_excel_report(mutations_df, output_file):
    """Generate a formatted Excel report with mutation data."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Mutation Summary"
        
        #TODO: vllt wieder rein aber macht nur probleme
        # Add title
        #ws['A1'] = "Sequence Alignment Mutation Report"
        #ws['A1'].font = Font(size=14, bold=True)
        #ws.merge_cells('A1:J1')
        #ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add headers with formatting
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        header_font = Font(bold=True)
        header_border = Border(bottom=Side(style='medium'))
        
        headers = ['Sample', 'Orientation', 'Nucleotide Pos', 'Original Codon', 
                   'Mutated Codon', 'AA Pos', 'Original AA', 'Mutated AA', 
                   'Silent?', 'Mutation Type']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
        
        # Add data
        current_sample = None
        row_offset = 0
        
        sorted_mutations = mutations_df.sort_values(['sample', 'nucleotide_position'])
        
        for r_idx, row in enumerate(dataframe_to_rows(sorted_mutations, index=False, header=False), 4):
            actual_row = r_idx + row_offset
            
            # Check if we're starting a new sample
            if current_sample is not None and row[0] != current_sample:
                # Insert blank row
                for c_idx in range(1, len(headers) + 1):
                    ws.cell(row=actual_row, column=c_idx, value="")
                row_offset += 1
                actual_row += 1
            
            current_sample = row[0]
            
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=actual_row, column=c_idx, value=value)
                
                # Highlight silent mutations in green, missense in yellow
                if c_idx == 9 and value == True:  # Silent column
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                elif c_idx == 10 and value == "Missense":  # Mutation type column
                    cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        # Add summary sheet
        summary_ws = wb.create_sheet(title="Summary Statistics")
        
        # Sample count
        sample_count = mutations_df['sample'].nunique()
        summary_ws['A1'] = "Total Samples Analyzed:"
        summary_ws['B1'] = sample_count
        
        # Mutation counts
        total_mutations = len(mutations_df)
        silent_mutations = mutations_df['is_silent'].sum()
        missense_mutations = total_mutations - silent_mutations
        
        summary_ws['A3'] = "Mutation Statistics"
        summary_ws['A3'].font = Font(bold=True)
        
        summary_ws['A4'] = "Total Mutations:"
        summary_ws['B4'] = total_mutations
        
        summary_ws['A5'] = "Silent Mutations:"
        summary_ws['B5'] = silent_mutations
        
        summary_ws['A6'] = "Missense Mutations:"
        summary_ws['B6'] = missense_mutations
        
        # Add per-sample statistics
        summary_ws['A8'] = "Mutations per Sample"
        summary_ws['A8'].font = Font(bold=True)
        
        sample_stats = mutations_df.groupby('sample').size().reset_index()
        sample_stats.columns = ['Sample', 'Mutation Count']
        
        for r_idx, row in enumerate(dataframe_to_rows(sample_stats, index=False), 9):
            for c_idx, value in enumerate(row, 1):
                summary_ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Add codon mutation analysis sheet
        try:
            codon_ws = wb.create_sheet(title="Codon Analysis")
            print("Creating Codon Analysis sheet...")
            
            # Group mutations by codon position and analyze frequency
            codon_mutations = mutations_df.groupby(['nucleotide_position', 'original_codon', 'mutated_codon']).agg({
                'sample': ['count', lambda x: ', '.join(sorted(set(x)))],
                'original_aa': 'first',
                'mutated_aa': 'first',
                'is_silent': 'first',
                'aa_position': 'first'  # Add amino acid position
            }).reset_index()
            
            codon_mutations.columns = ['Position', 'Original Codon', 'Mutated Codon', 
                                      'Occurrence Count', 'Samples', 'Original AA', 
                                      'Mutated AA', 'Silent', 'AA Pos']
            
            # Sort by position and occurrence count
            codon_mutations = codon_mutations.sort_values(['Position', 'Occurrence Count'], ascending=[True, False])
            
            # Add headers
            codon_headers = ['Position', 'Original Codon', 'Mutated Codon', 'Occurrence Count', 
                            'AA Pos', 'Original AA', 'Mutated AA', 'Silent', 'Samples']
            
            for col, header in enumerate(codon_headers, 1):
                cell = codon_ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                cell.border = Border(bottom=Side(style='medium'))
            
            # Reorder columns to put AA Pos before Original AA and Samples at the end
            codon_mutations = codon_mutations[['Position', 'Original Codon', 'Mutated Codon', 
                                              'Occurrence Count', 'AA Pos', 'Original AA', 
                                              'Mutated AA', 'Silent', 'Samples']]
            
            # Add data
            for r_idx, row in enumerate(dataframe_to_rows(codon_mutations, index=False, header=False), 2):
                for c_idx, value in enumerate(row, 1):
                    cell = codon_ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Highlight silent mutations in green
                    if c_idx == 8 and value == True:  # Silent column (now at position 8)
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            print("Codon Analysis sheet completed")
        except Exception as e:
            print(f"Error creating Codon Analysis sheet: {e}")
        
        # Add variant analysis sheet
        try:
            variant_ws = wb.create_sheet(title="Variant Analysis")
            print("Creating Variant Analysis sheet...")
            
            # Identify unique mutation patterns (variants)
            # Convert mutations to a format that can be used for grouping
            def get_mutation_signature(group):
                return frozenset([(row['nucleotide_position'], row['original_codon'], row['mutated_codon']) 
                                 for _, row in group.iterrows()])
            
            # Group samples by their mutation patterns
            sample_groups = {}
            for sample, group in mutations_df.groupby('sample'):
                signature = get_mutation_signature(group)
                if signature in sample_groups:
                    sample_groups[signature].append(sample)
                else:
                    sample_groups[signature] = [sample]
            
            # Create variant data for the sheet
            variant_data = []
            for i, (signature, samples) in enumerate(sample_groups.items(), 1):
                #variant_name = f"Variant {i}"
                variant_name = f"{samples[0]}"

                #print(i)
                #print(signature)
                #print(samples)
                sample_count = len(samples)
                sample_list = ", ".join(sorted(samples))
                
                # Get mutation details for this variant
                mutation_details = []
                # Convert signature to list and sort by position (first element of each tuple)
                sorted_signature = sorted(signature, key=lambda x: x[0])
                for pos, orig, mut in sorted_signature:
                    mutation_row = mutations_df[(mutations_df['nucleotide_position'] == pos) & 
                                               (mutations_df['original_codon'] == orig) & 
                                               (mutations_df['mutated_codon'] == mut)].iloc[0]
                    mutation_details.append(f"{pos}: {orig}->{mut} ({mutation_row['original_aa']}->{mutation_row['mutated_aa']})")
                
                mutation_summary = "; ".join(mutation_details)
                
                variant_data.append([variant_name, sample_count, sample_list, len(signature), mutation_summary])
            
            # Sort variants by frequency
            variant_data.sort(key=lambda x: x[1], reverse=True)
            
            # Add headers
            variant_headers = ['Variant', 'Frequency', 'Mutation Count', 'Mutations', 'Samples']
            
            for col, header in enumerate(variant_headers, 1):
                cell = variant_ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                cell.border = Border(bottom=Side(style='medium'))
            
            # Reorder variant data to put Samples at the end
            reordered_variant_data = []
            for row in variant_data:
                # Original order: [variant_name, sample_count, sample_list, mutation_count, mutation_summary]
                # New order: [variant_name, sample_count, mutation_count, mutation_summary, sample_list]
                reordered_variant_data.append([row[0], row[1], row[3], row[4], row[2]])
            
            # Add data
            for r_idx, row in enumerate(reordered_variant_data, 2):
                for c_idx, value in enumerate(row, 1):
                    variant_ws.cell(row=r_idx, column=c_idx, value=value)
            print("Variant Analysis sheet completed")
        except Exception as e:
            print(f"Error creating Variant Analysis sheet: {e}")
        
        # Before saving
        print(f"Workbook has {len(wb.worksheets)} sheets: {[ws.title for ws in wb.worksheets]}")
        
        # Auto-adjust column widths for all sheets
        for sheet in wb.worksheets:
            column_widths = {}
            for row in sheet.rows:
                for i, cell in enumerate(row):
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.value:
                        # Get the column letter from the cell's coordinate
                        column_letter = cell.column_letter
                        # Calculate the length of the cell value
                        try:
                            if sheet.title == "Variant Analysis" and i == 4:  # Mutations column (0-based index)
                                cell_length = min(len(str(cell.value)), 100)
                            else:
                                cell_length = len(str(cell.value))
                            
                            # Update the width if this cell is wider than what we've seen so far
                            current_width = column_widths.get(column_letter, 0)
                            column_widths[column_letter] = max(current_width, cell_length)
                        except:
                            pass
            
            # Set column widths with some padding
            for col_letter, width in column_widths.items():
                sheet.column_dimensions[col_letter].width = width + 2
        
        # Save workbook
        wb.save(output_file)
        return output_file
    except Exception as e:
        print(f"Error generating Excel report: {e}")
        return None

def generate_mutation_plots(mutations_df, output_dir):
    """Generate plots visualizing the mutation data."""
    os.makedirs(output_dir, exist_ok=True)
    
    # 1. Mutation distribution by sample
    plt.figure(figsize=(10, 6))
    sample_counts = mutations_df.groupby('sample').size().sort_values(ascending=False)
    sns.barplot(x=sample_counts.index, y=sample_counts.values, width=0.5)
    plt.title('Mutations per Sample')
    plt.xlabel('Sample')
    plt.ylabel('Number of Mutations')
    plt.xticks(rotation=45, ha='right')
    
    # Format y-axis to show only integers
    from matplotlib.ticker import MaxNLocator
    plt.gca().yaxis.set_major_locator(MaxNLocator(integer=True))
    
    plt.tight_layout()
    plt.savefig(os.path.join(output_dir, 'mutations_per_sample.png'))
    plt.close()
    
    # 2. Silent vs Missense mutations
    plt.figure(figsize=(8, 8))
    mutation_types = mutations_df['mutation_type'].value_counts()
    plt.pie(mutation_types, labels=mutation_types.index, autopct='%1.1f%%', 
            colors=['#E2EFDA', '#FCE4D6'])
    plt.title('Distribution of Mutation Types')
    plt.savefig(os.path.join(output_dir, 'mutation_types.png'))
    plt.close()
    
    # 3. Mutation positions along the sequence
    plt.figure(figsize=(12, 6))
    sns.histplot(data=mutations_df, x='nucleotide_position', bins=30, binwidth=0.5, color='#6699CC')
    plt.title('Distribution of Mutations Along the Sequence')
    plt.xlabel('Nucleotide Position')
    plt.ylabel('Number of Mutations')
    plt.savefig(os.path.join(output_dir, 'mutation_positions.png'))
    plt.close()
    
    return output_dir

def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='Generate reports from mutation data')
    parser.add_argument('--input', required=True, help='Input Excel file with mutation data')
    parser.add_argument('--output', default='mutation_report.xlsx', help='Output Excel report file')
    parser.add_argument('--plots', default='plots', help='Directory for output plots')
    
    args = parser.parse_args()
    
    # Load mutation data
    mutations_df = pd.read_excel(args.input)
    
    # Generate Excel report
    report_file = generate_excel_report(mutations_df, args.output)
    print(f"Excel report generated: {report_file}")
    
    # Generate plots
    plots_dir = generate_mutation_plots(mutations_df, args.plots)
    print(f"Plots generated in directory: {plots_dir}")

if __name__ == "__main__":
    main()